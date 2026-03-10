"""SQLite-based cache for extraction code reuse."""

from __future__ import annotations

import hashlib
import sqlite3
from pathlib import Path
from typing import Optional

from bds.config import CacheConfig


class MappingCache:
    """Cache extraction code keyed by file structure signatures.

    Same-structure files (e.g., NASA B0005.mat and B0006.mat) share the same
    cached extraction code.
    """

    def __init__(self, config: CacheConfig):
        self.config = config
        self._conn: Optional[sqlite3.Connection] = None
        if config.enabled:
            self._init_db()

    def _init_db(self) -> None:
        db_path = Path(self.config.db_path)
        db_path.parent.mkdir(parents=True, exist_ok=True)
        self._conn = sqlite3.connect(str(db_path))
        self._conn.execute(
            """
            CREATE TABLE IF NOT EXISTS cache (
                signature TEXT PRIMARY KEY,
                code TEXT NOT NULL,
                source_file TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        self._conn.commit()

    def get(self, file_path: Path) -> Optional[str]:
        """Look up cached extraction code by file structure signature."""
        if not self.config.enabled or self._conn is None:
            return None
        sig = self._compute_signature(file_path)
        row = self._conn.execute(
            "SELECT code FROM cache WHERE signature = ?", (sig,)
        ).fetchone()
        return row[0] if row else None

    def save(self, file_path: Path, code: str) -> None:
        """Save extraction code for a file's structure signature."""
        if not self.config.enabled or self._conn is None:
            return
        sig = self._compute_signature(file_path)
        self._conn.execute(
            "INSERT OR REPLACE INTO cache (signature, code, source_file) VALUES (?, ?, ?)",
            (sig, code, str(file_path)),
        )
        self._conn.commit()

    def list_entries(self) -> list[dict]:
        """List all cached entries."""
        if not self.config.enabled or self._conn is None:
            return []
        rows = self._conn.execute(
            "SELECT signature, source_file, created_at FROM cache ORDER BY created_at DESC"
        ).fetchall()
        return [
            {"signature": r[0], "source_file": r[1], "created_at": r[2]}
            for r in rows
        ]

    def clear(self) -> int:
        """Clear all cached entries. Returns number of entries deleted."""
        if not self.config.enabled or self._conn is None:
            return 0
        cursor = self._conn.execute("DELETE FROM cache")
        self._conn.commit()
        return cursor.rowcount

    def _compute_signature(self, file_path: Path) -> str:
        """Compute a structural signature of the file.

        For CSV: hash of the header row.
        For MAT: hash of top-level keys + structure.
        For XLSX: hash of sheet names + each sheet's header row.
        Generic fallback: hash of first 4KB.
        """
        path = Path(file_path)
        ext = path.suffix.lower()

        try:
            if ext in (".csv", ".tsv", ".txt"):
                return self._sig_csv(path)
            elif ext in (".xlsx", ".xls"):
                return self._sig_excel(path)
            elif ext == ".mat":
                return self._sig_mat(path)
            elif ext == ".json":
                return self._sig_json(path)
            elif ext in (".h5", ".hdf5"):
                return self._sig_hdf5(path)
            elif ext in (".pkl", ".pickle"):
                return self._sig_pickle(path)
            else:
                return self._sig_generic(path)
        except Exception:
            return self._sig_generic(path)

    @staticmethod
    def _sig_csv(path: Path) -> str:
        with open(path, errors="replace") as f:
            header = f.readline()
        return _hash(f"csv:{header.strip()}")

    @staticmethod
    def _sig_excel(path: Path) -> str:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        parts = []
        for name in wb.sheetnames:
            ws = wb[name]
            header_row = next(ws.iter_rows(max_row=1, values_only=True), ())
            parts.append(f"{name}:{header_row}")
        wb.close()
        return _hash("xlsx:" + "|".join(parts))

    @staticmethod
    def _sig_mat(path: Path) -> str:
        try:
            import h5py
            with h5py.File(path, "r") as f:
                keys = sorted(f.keys())
                return _hash("mat_h5:" + ",".join(keys))
        except Exception:
            import scipy.io
            data = scipy.io.loadmat(path, squeeze_me=False)
            keys = sorted(k for k in data.keys() if not k.startswith("__"))
            return _hash("mat_v5:" + ",".join(keys))

    @staticmethod
    def _sig_json(path: Path) -> str:
        import json
        raw = path.read_text()[:10000]
        data = json.loads(raw)
        if isinstance(data, dict):
            return _hash("json:" + ",".join(sorted(data.keys())))
        elif isinstance(data, list) and data:
            if isinstance(data[0], dict):
                return _hash("json_arr:" + ",".join(sorted(data[0].keys())))
        return _hash("json:" + raw[:500])

    @staticmethod
    def _sig_hdf5(path: Path) -> str:
        import h5py
        with h5py.File(path, "r") as f:
            keys = sorted(f.keys())
        return _hash("hdf5:" + ",".join(keys))

    @staticmethod
    def _sig_pickle(path: Path) -> str:
        import pickle
        with open(path, "rb") as f:
            data = pickle.load(f)
        if isinstance(data, dict):
            return _hash("pkl:" + ",".join(sorted(str(k) for k in data.keys())))
        return _hash(f"pkl:{type(data).__name__}")

    @staticmethod
    def _sig_generic(path: Path) -> str:
        chunk = path.read_bytes()[:4096]
        return _hash(f"generic:{chunk!r}")


def _hash(content: str) -> str:
    return hashlib.sha256(content.encode()).hexdigest()[:16]

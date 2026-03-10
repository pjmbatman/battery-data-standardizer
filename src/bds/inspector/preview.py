"""File structure preview — converts any file into a human/LLM-readable text summary."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any


def inspect_file(path: Path) -> str:
    """Return a text preview of the file's structure and sample data."""
    path = Path(path)
    if not path.exists():
        return f"[File not found: {path}]"

    ext = path.suffix.lower()
    size = path.stat().st_size

    try:
        if ext in (".csv", ".tsv", ".txt"):
            return _preview_text(path, size)
        elif ext in (".xlsx", ".xls"):
            return _preview_excel(path, size)
        elif ext == ".mat":
            return _preview_mat(path, size)
        elif ext == ".json":
            return _preview_json(path, size)
        elif ext in (".pkl", ".pickle"):
            return _preview_pickle(path, size)
        elif ext in (".h5", ".hdf5"):
            return _preview_hdf5(path, size)
        else:
            return _preview_unknown(path, ext, size)
    except Exception as exc:
        return f"[Error inspecting {path.name}: {exc}]"


# ---------------------------------------------------------------------------
# Format-specific preview helpers
# ---------------------------------------------------------------------------

def _preview_text(path: Path, size: int) -> str:
    lines = path.read_text(errors="replace").splitlines()[:20]
    total_lines = sum(1 for _ in open(path, errors="replace"))
    header = f"[CSV/Text file, {size:,} bytes, ~{total_lines} lines]"
    return header + "\n" + "\n".join(lines)


def _preview_excel(path: Path, size: int) -> str:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result = f"[Excel file, {size:,} bytes, sheets: {wb.sheetnames}]\n"
    for name in wb.sheetnames[:5]:
        ws = wb[name]
        rows = list(ws.iter_rows(max_row=5, values_only=True))
        result += f"\n--- Sheet: {name} ---\n"
        for row in rows:
            result += str(row) + "\n"
    wb.close()
    return result


def _preview_mat(path: Path, size: int) -> str:
    # Try HDF5 (v7.3) first, then scipy (v5)
    try:
        import h5py

        with h5py.File(path, "r") as f:
            return f"[MAT v7.3 (HDF5), {size:,} bytes]\n" + _h5py_tree(f)
    except Exception:
        pass

    import scipy.io

    data_raw = scipy.io.loadmat(path, squeeze_me=False)
    tree = _scipy_tree(data_raw, max_depth=6)
    # Load again with squeeze for easier nested traversal
    data_squeezed = scipy.io.loadmat(path, squeeze_me=True)
    sample = _mat_v5_sample(data_squeezed)
    # Generate concrete access code example
    access_code = _mat_v5_access_example(data_squeezed)
    hint = (
        "\n\n## Access hint\n"
        "Use scipy.io.loadmat(path, squeeze_me=True) for easier access.\n"
        "Do NOT use h5py for MAT v5 files.\n"
        "IMPORTANT: With squeeze_me=True, shape=() arrays need [()] to extract the value.\n"
        "Example: data['key']['field'][()] to get the inner array from a 0-d wrapper."
    )
    return f"[MAT v5, {size:,} bytes]\n{tree}{sample}{hint}{access_code}"


def _preview_json(path: Path, size: int) -> str:
    raw = path.read_text(errors="replace")[:50_000]
    data = json.loads(raw)
    truncated = _truncate_structure(data)
    return f"[JSON file, {size:,} bytes]\n" + json.dumps(truncated, indent=2, default=str)


def _preview_pickle(path: Path, size: int) -> str:
    import pickle

    with open(path, "rb") as f:
        data = pickle.load(f)
    return f"[Pickle file, {size:,} bytes, type: {type(data).__name__}]\n" + _describe_object(data)


def _preview_hdf5(path: Path, size: int) -> str:
    import h5py

    with h5py.File(path, "r") as f:
        return f"[HDF5 file, {size:,} bytes]\n" + _h5py_tree(f)


def _preview_unknown(path: Path, ext: str, size: int) -> str:
    try:
        content = path.read_text(errors="replace")[:5000]
        return f"[Unknown format: {ext}, {size:,} bytes]\n" + content
    except Exception:
        return f"[Binary file: {ext}, {size:,} bytes]"


# ---------------------------------------------------------------------------
# Helper utilities
# ---------------------------------------------------------------------------

def _h5py_tree(group, prefix: str = "", max_depth: int = 4, _depth: int = 0) -> str:
    """Recursively render an HDF5 group as a text tree."""
    import h5py

    if _depth >= max_depth:
        return prefix + "...\n"

    lines: list[str] = []
    items = list(group.keys())[:50]
    for key in items:
        item = group[key]
        if isinstance(item, h5py.Group):
            lines.append(f"{prefix}{key}/ (group, {len(item)} items)")
            lines.append(_h5py_tree(item, prefix + "  ", max_depth, _depth + 1))
        elif isinstance(item, h5py.Dataset):
            lines.append(f"{prefix}{key}: shape={item.shape}, dtype={item.dtype}")
        else:
            lines.append(f"{prefix}{key}: {type(item).__name__}")
    if len(group.keys()) > 50:
        lines.append(f"{prefix}... ({len(group.keys()) - 50} more items)")
    return "\n".join(lines)


def _scipy_tree(data: dict, prefix: str = "", max_depth: int = 6, _depth: int = 0) -> str:
    """Recursively describe a scipy loadmat result."""
    import numpy as np

    if _depth >= max_depth:
        return prefix + "...\n"

    lines: list[str] = []
    for key in sorted(data.keys()):
        if key.startswith("__"):
            continue
        val = data[key]
        if isinstance(val, np.ndarray):
            if val.dtype.names:
                lines.append(f"{prefix}{key}: structured array, shape={val.shape}, fields={list(val.dtype.names)}")
                if val.size > 0 and _depth < max_depth - 1:
                    sample = val.flat[0]
                    for fname in val.dtype.names[:10]:
                        fval = sample[fname]
                        if isinstance(fval, np.ndarray):
                            lines.append(f"{prefix}  .{fname}: shape={fval.shape}, dtype={fval.dtype}")
                        else:
                            lines.append(f"{prefix}  .{fname}: {type(fval).__name__} = {_safe_repr(fval)}")
            elif val.dtype == object:
                lines.append(f"{prefix}{key}: object array, shape={val.shape}")
                if val.size > 0 and _depth < max_depth - 1:
                    inner = val.flat[0]
                    if isinstance(inner, np.ndarray) and inner.dtype.names:
                        lines.append(
                            _scipy_tree(
                                {n: inner[n] for n in inner.dtype.names},
                                prefix + "  ", max_depth, _depth + 1,
                            )
                        )
                    else:
                        lines.append(f"{prefix}  [0]: {type(inner).__name__}, {_safe_repr(inner)}")
            else:
                lines.append(f"{prefix}{key}: ndarray, shape={val.shape}, dtype={val.dtype}")
                if val.size <= 5:
                    lines.append(f"{prefix}  values: {val.tolist()}")
        elif isinstance(val, dict):
            lines.append(f"{prefix}{key}: dict ({len(val)} keys)")
            lines.append(_scipy_tree(val, prefix + "  ", max_depth, _depth + 1))
        else:
            lines.append(f"{prefix}{key}: {type(val).__name__} = {_safe_repr(val)}")
    return "\n".join(lines)


def _mat_v5_sample(data: dict) -> str:
    """Extract sample data from a MAT v5 structured array — especially NASA-style nested data."""
    import numpy as np

    # Re-load with squeeze_me=True for easier traversal
    lines: list[str] = []
    for key in sorted(data.keys()):
        if key.startswith("__"):
            continue
        val = data[key]
        if not isinstance(val, np.ndarray) or not val.dtype.names:
            continue

        try:
            sample = val.flat[0]
            for fname in val.dtype.names:
                fval = sample[fname]
                # Handle object arrays or scalar ndarray of structured type
                if isinstance(fval, np.ndarray):
                    # Unwrap scalar arrays
                    arr = fval[()] if fval.shape == () else fval
                    if not hasattr(arr, '__len__') or not len(arr):
                        continue

                    # Check if elements are structured records (NASA pattern)
                    first_elem = arr.flat[0] if hasattr(arr, 'flat') else arr[0] if hasattr(arr, '__getitem__') else None
                    if first_elem is None:
                        continue

                    if isinstance(first_elem, np.void) and first_elem.dtype.names:
                        # arr is an array of structured records (cycles)
                        # Show types/categories if there's a 'type' field
                        types_seen = {}
                        for i, elem in enumerate(arr):
                            t = str(elem['type']) if 'type' in (elem.dtype.names or []) else f"idx{i}"
                            if t not in types_seen:
                                types_seen[t] = i
                            if len(types_seen) >= 5:
                                break

                        lines.append(f"\n## {key}.{fname}: {len(arr)} records")
                        if 'type' in (first_elem.dtype.names or []):
                            from collections import Counter
                            type_counts = Counter(str(e['type']) for e in arr)
                            lines.append(f"Types: {dict(type_counts)}")

                        # Show data fields — prioritize charge/discharge over other types
                        priority_types = ['charge', 'discharge']
                        shown = set()
                        for t in priority_types + list(types_seen.keys()):
                            if t in shown or t not in types_seen:
                                continue
                            shown.add(t)
                            if len(shown) > 3:  # Limit to 3 types
                                break
                            idx = types_seen[t]
                            elem = arr[idx]
                            lines.append(f"\n  [{t}] record fields: {list(elem.dtype.names)}")
                            if 'data' in (elem.dtype.names or []):
                                dval = elem['data']
                                if isinstance(dval, np.ndarray) and dval.shape == () and dval.dtype.names:
                                    dval = dval[()]
                                if isinstance(dval, np.void) and dval.dtype.names:
                                    lines.append(f"    .data fields: {list(dval.dtype.names)}")
                                    for sfn in dval.dtype.names[:7]:
                                        sfv = dval[sfn]
                                        if isinstance(sfv, np.ndarray) and sfv.dtype.kind == 'f':
                                            sv = sfv.flat[:2].tolist() if sfv.size > 0 else []
                                            lines.append(f"      {sfn}: shape={sfv.shape}, sample={sv}")
                                        elif not isinstance(sfv, np.ndarray):
                                            lines.append(f"      {sfn}: {_safe_repr(sfv)[:60]}")
        except Exception:
            pass
    if lines:
        return "\n" + "\n".join(lines)
    return ""


def _mat_v5_access_example(data: dict) -> str:
    """Generate a concrete Python code snippet showing how to access the MAT v5 data."""
    import numpy as np

    lines: list[str] = []
    for key in sorted(data.keys()):
        if key.startswith("__"):
            continue
        val = data[key]
        if not isinstance(val, np.ndarray) or not val.dtype.names:
            continue
        if 'cycle' not in val.dtype.names:
            continue

        # This looks like a battery dataset with cycles
        try:
            cycle_field = val['cycle']
            # Handle 0-d array
            if cycle_field.shape == ():
                cycle_arr = cycle_field[()]
            else:
                cycle_arr = cycle_field

            if not hasattr(cycle_arr, '__len__') or len(cycle_arr) == 0:
                continue

            first = cycle_arr[0]
            if not hasattr(first, 'dtype') or not first.dtype.names:
                continue

            lines.append(f"\n\n## Working code example for this file:")
            lines.append("```python")
            lines.append("import scipy.io")
            lines.append(f"data = scipy.io.loadmat(FILE_PATH, squeeze_me=True)")
            lines.append(f"cell = data['{key}']")
            lines.append(f"cycles_raw = cell['cycle'][()]  # [()] to unwrap 0-d array")
            lines.append(f"# cycles_raw is an array of {len(cycle_arr)} records")

            # Show how to access fields
            if 'type' in first.dtype.names:
                lines.append(f"# Each record has fields: {list(first.dtype.names)}")
                lines.append("# Filter by type: [c for c in cycles_raw if str(c['type']) in ('charge', 'discharge')]")

            if 'data' in first.dtype.names:
                # Find a charge or discharge record
                for c in cycle_arr:
                    ctype = str(c['type']) if 'type' in c.dtype.names else ''
                    if ctype in ('charge', 'discharge'):
                        dval = c['data']
                        if isinstance(dval, np.ndarray) and dval.shape == ():
                            dval = dval[()]
                        if hasattr(dval, 'dtype') and dval.dtype.names:
                            lines.append(f"# For a {ctype} cycle: c['data'][()] has fields {list(dval.dtype.names)}")
                            lines.append(f"# Access: c['data'][()]['Voltage_measured'] → array of floats")
                        break
            lines.append("```")
        except Exception:
            pass

    return "\n".join(lines) if lines else ""


def _describe_object(obj: Any, prefix: str = "", max_depth: int = 3, _depth: int = 0) -> str:
    """Recursively describe an arbitrary Python object."""
    if _depth >= max_depth:
        return prefix + "..."

    lines: list[str] = []

    if isinstance(obj, dict):
        lines.append(f"{prefix}dict ({len(obj)} keys)")
        for i, (k, v) in enumerate(obj.items()):
            if i >= 20:
                lines.append(f"{prefix}  ... ({len(obj) - 20} more keys)")
                break
            lines.append(f"{prefix}  {k}: {_type_summary(v)}")
            if isinstance(v, (dict, list)) and _depth < max_depth - 1:
                lines.append(_describe_object(v, prefix + "    ", max_depth, _depth + 1))
    elif isinstance(obj, (list, tuple)):
        tname = type(obj).__name__
        lines.append(f"{prefix}{tname} (len={len(obj)})")
        for i, item in enumerate(obj[:3]):
            lines.append(f"{prefix}  [{i}]: {_type_summary(item)}")
            if isinstance(item, (dict, list)) and _depth < max_depth - 1:
                lines.append(_describe_object(item, prefix + "    ", max_depth, _depth + 1))
        if len(obj) > 3:
            lines.append(f"{prefix}  ... ({len(obj) - 3} more)")
    else:
        lines.append(f"{prefix}{_type_summary(obj)}")

    return "\n".join(lines)


def _truncate_structure(data: Any, max_items: int = 3, max_depth: int = 4, _depth: int = 0) -> Any:
    """Keep JSON structure but truncate long arrays/objects."""
    if _depth >= max_depth:
        return "..."

    if isinstance(data, dict):
        items = list(data.items())[:max_items]
        result = {k: _truncate_structure(v, max_items, max_depth, _depth + 1) for k, v in items}
        if len(data) > max_items:
            result[f"... ({len(data) - max_items} more keys)"] = "..."
        return result
    elif isinstance(data, list):
        items = [_truncate_structure(v, max_items, max_depth, _depth + 1) for v in data[:max_items]]
        if len(data) > max_items:
            items.append(f"... ({len(data) - max_items} more items)")
        return items
    else:
        return data


def _type_summary(obj: Any) -> str:
    """One-line type summary of an object."""
    import numpy as np

    if isinstance(obj, np.ndarray):
        return f"ndarray(shape={obj.shape}, dtype={obj.dtype})"
    if isinstance(obj, dict):
        return f"dict({len(obj)} keys)"
    if isinstance(obj, (list, tuple)):
        return f"{type(obj).__name__}(len={len(obj)})"
    if isinstance(obj, str):
        return f"str(len={len(obj)}): {obj[:80]!r}"
    return f"{type(obj).__name__}: {_safe_repr(obj)}"


def _safe_repr(obj: Any, max_len: int = 200) -> str:
    """Safe repr that won't explode on large objects."""
    try:
        r = repr(obj)
        if len(r) > max_len:
            return r[:max_len] + "..."
        return r
    except Exception:
        return f"<{type(obj).__name__}>"
